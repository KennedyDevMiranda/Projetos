import os
from pathlib import Path
import sys
from datetime import datetime
from openpyxl import load_workbook

# =========================
# CONFIGURAÇÕES
# =========================

#PASTA_SAIDA = r"C:\Users\kenne\Documents"
PASTA_SAIDA = r"C:\Users\micro\Documents"

NOME_ABA = None
LINHA_CABECALHO = 1

# Se True, remove o cabeçalho no final
APAGAR_LINHA_1_NO_FINAL = True 

# Colunas necessárias para a lógica e para o arquivo final
COLUNAS_PARA_MANTER = [
    "Nome Aluno",
    #"Email Aluno",
    #"Celular", # Coluna do Aluno
    #"Responsável Legal",
    #"Celular Responsável Financeiro", # Coluna do Responsável
]

def garantir_pasta_saida():
    global PASTA_SAIDA
    try:
        os.makedirs(PASTA_SAIDA, exist_ok=True)
    except PermissionError:
        fallback = str(Path.home() / "Documents" / "Excel_Limpo")
        PASTA_SAIDA = fallback
        os.makedirs(PASTA_SAIDA, exist_ok=True)

def limpar_numeros_duplicados(ws, linha_cabecalho=1):
    """
    Se o celular do aluno for igual ao do responsável, 
    apaga o celular do aluno para evitar duplicidade no envio.
    """
    # 1. Mapear onde estão as colunas de celular
    col_celular_aluno = None
    col_celular_resp = None

    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=linha_cabecalho, column=col).value
        if header == "Celular":
            col_celular_aluno = col
        elif header == "Celular Responsável Financeiro":
            col_celular_resp = col

    if not col_celular_aluno or not col_celular_resp:
        print("[Aviso] Colunas de celular não encontradas para desduplicação.")
        return

    # 2. Percorrer as linhas (pulando o cabeçalho)
    for row in range(linha_cabecalho + 1, ws.max_row + 1):
        cel_aluno = str(ws.cell(row=row, column=col_celular_aluno).value).strip()
        cel_resp = str(ws.cell(row=row, column=col_celular_resp).value).strip()

        # Se os números forem iguais (e não estiverem vazios)
        if cel_aluno == cel_resp and cel_aluno not in ["None", ""]:
            # Deixamos o do Responsável e limpamos o do Aluno
            ws.cell(row=row, column=col_celular_aluno).value = None

def manter_apenas_colunas(ws, nomes_para_manter, linha_cabecalho=1):
    colunas = []
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=linha_cabecalho, column=col).value
        nome = "" if valor is None else str(valor).strip()
        colunas.append((col, nome))

    indices_para_apagar = []
    for idx, nome in colunas:
        if nome not in nomes_para_manter:
            indices_para_apagar.append(idx)

    for idx in sorted(indices_para_apagar, reverse=True):
        ws.delete_cols(idx)

def main():
    global PASTA_SAIDA

    if len(sys.argv) < 2:
        print("Uso: python limpar_colunas_excel.py \"arquivo.xlsx\"")
        return

    caminho_entrada = sys.argv[1]
    if not os.path.exists(caminho_entrada):
        print("Arquivo não encontrado:", caminho_entrada)
        return

    garantir_pasta_saida()
    wb = load_workbook(caminho_entrada)
    ws = wb[NOME_ABA] if NOME_ABA else wb.worksheets[0]

    # 1. Primeiro filtramos as colunas
    manter_apenas_colunas(ws, COLUNAS_PARA_MANTER, LINHA_CABECALHO)
    
    # 2. AGORA aplicamos a lógica de desduplicação de números
    limpar_numeros_duplicados(ws, LINHA_CABECALHO)

    # 3. Opcional: apagar linha 1 (cabeçalho)
    if APAGAR_LINHA_1_NO_FINAL:
        ws.delete_rows(1)

    nome_base = os.path.splitext(os.path.basename(caminho_entrada))[0]
    datahora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    caminho_saida = os.path.join(PASTA_SAIDA, f"{nome_base}_ALUNOS_LIMPO_{datahora}.xlsx")

    wb.save(caminho_saida)
    print(f"✅ Arquivo processado com sucesso!")
    print(f"📍 Salvo em: {caminho_saida}")

if __name__ == "__main__":
    main()