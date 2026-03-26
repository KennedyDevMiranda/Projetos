import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

# (Opcional) Janela para escolher o arquivo
try:
    from tkinter import Tk
    from tkinter.filedialog import askopenfilename
    USA_JANELA = True
except ImportError:
    USA_JANELA = False


# =========================
# CONFIGURAÇÕES (edite aqui)
# =========================

# Pasta onde vai salvar o arquivo final (mude para o seu caminho)
#PASTA_SAIDA = r"C:\Users\kenne\Documents"
PASTA_SAIDA = r"C:\Users\micro\Documents"

# Nome da aba (sheet). Se quiser a primeira aba, deixe None.
NOME_ABA = None  # exemplo: "Planilha1"

# Colunas que você quer APAGAR (pelos nomes do cabeçalho, linha 1)
COLUNAS_PARA_APAGAR = [
    "Contrato",
    "Data Inicial",
    "Data Final",
    "Status",
    "Telefone Aluno",
    "Data Limite",
    "Aulas",
    "Quantidade Agendamentos",
    "Data Limite Próxima Matéria",
    "Data Início",
    "Quantidade Matérias Restantes",
    "Data Final Prevista",
    "Recebimentos Futuros",
    "Tipo Contrato",
    "Nome Responsável Financeiro",
    "Tel. Residencial Responsável Financeiro",
    "Telefone Celular Responsável Financeiro",
]

# Linha onde ficam os cabeçalhos (normalmente 1)
LINHA_CABECALHO = 1


def escolher_arquivo():
    """Abre uma janela para selecionar o arquivo Excel."""
    if not USA_JANELA:
        raise RuntimeError("Tkinter não está disponível. Defina o caminho do arquivo manualmente.")

    Tk().withdraw()
    caminho = askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Excel files", "*.xlsx")],
    )
    return caminho

def garantir_pasta_saida():
    global PASTA_SAIDA

    try:
        os.makedirs(PASTA_SAIDA, exist_ok=True)
        return
    except PermissionError:
        # Fallback: Documentos do usuário
        fallback = str(Path.home() / "Documents" / "Excel_Limpo")
        PASTA_SAIDA = fallback
        os.makedirs(PASTA_SAIDA, exist_ok=True)

def apagar_colunas_por_nome(ws, nomes_colunas, linha_cabecalho=1):
    """
    Remove colunas com base nos nomes do cabeçalho.
    Importante: deletar da direita para a esquerda para não bagunçar os índices.
    """
    # Mapeia: nome_do_cabecalho -> indice_da_coluna (1-based)
    mapa = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=linha_cabecalho, column=col).value
        if valor is None:
            continue
        nome = str(valor).strip()
        mapa[nome] = col

    # Descobre quais índices existem
    indices_para_apagar = []
    for nome in nomes_colunas:
        if nome in mapa:
            indices_para_apagar.append(mapa[nome])
        else:
            print(f"[Aviso] Coluna '{nome}' não encontrada no cabeçalho.")

    # Apaga da direita para a esquerda
    for idx in sorted(indices_para_apagar, reverse=True):
        ws.delete_cols(idx)


def main():
    garantir_pasta_saida()

    # Se veio do C# com argumentos:
    # python script.py "arquivo.xlsx" "pasta_saida"
    if len(sys.argv) >= 2:
        caminho_entrada = sys.argv[1]
    else:
        caminho_entrada = escolher_arquivo() if USA_JANELA else r"Documentos\arquivo.xlsx"

    if len(sys.argv) >= 3:
        global PASTA_SAIDA
        PASTA_SAIDA = sys.argv[2]
        garantir_pasta_saida()

    # 2) Abrir workbook
    wb = load_workbook(caminho_entrada)
    ws = wb[NOME_ABA] if NOME_ABA else wb.worksheets[0]

    # 3) Apagar colunas
    apagar_colunas_por_nome(ws, COLUNAS_PARA_APAGAR, LINHA_CABECALHO)

    # 3.1) Apagar a primeira linha (linha 1)
    ws.delete_rows(1)

    # 4) Salvar com nome automático
    nome_base = os.path.splitext(os.path.basename(caminho_entrada))[0]
    datahora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    caminho_saida = os.path.join(PASTA_SAIDA, f"{nome_base}_LIVROS_{datahora}.xlsx")

    wb.save(caminho_saida)
    print(f"✅ Arquivo salvo em: {caminho_saida}")


if __name__ == "__main__":
    main()